import requests
from termcolor import colored
from bs4 import BeautifulSoup
import openpyxl

# req = requests.get('https://www.ua-football.com/sport')
# html = req.content
# soup = BeautifulSoup(html, 'lxml')
#
# news = soup.find_all('li', class_='liga-news-item')
#
# results = []
#
# for item in news:
#     title = item.find('span', class_='d-block').get_text(strip=True)
#     print(colored(title, 'green'))
#     desc = item.find('span', class_='name-dop').get_text(strip=True)
#     print(colored(desc, 'red'))
#     href = item.a.get('href')
#     print(colored(href, 'blue'))
#     results.append({
#         'title': title,
#         'desc': desc,
#         'href': href,
#     })


def parse(url):
    req = requests.get(url)

    html = req.content

    soup = BeautifulSoup(html, 'lxml')
    news = soup.find_all('li', class_='liga-news-item')

    results = []

    for item in news:
        title = item.find('span', class_='d-block').get_text(strip=True)
        desc = item.find('span', class_='name-dop').get_text(strip=True)
        href = item.a.get('href')
        results.append({
            'title': title,
            'desc': desc,
            'href': href,
        })
    return results


def main():
    url = 'https://www.ua-football.com/sport?page={number}'
    i = 1
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])

    while i <= 5:
        news = parse(url.format(number=i))
        wb.create_sheet(title='Page {number}'.format(number=i))
        sheet = wb['Page {number}'.format(number=i)]
        for elem in news:
            cell = sheet.cell(row=news.index(elem) + 1, column=1)
            cell.value = elem['title']
            cell1 = sheet.cell(row=news.index(elem) + 1, column=2)
            cell1.value = elem['desc']
            cell2 = sheet.cell(row=news.index(elem) + 1, column=3)
            cell2.value = elem['href']
        i += 1

    wb.save('parse.xlsx')


if __name__ == '__main__':
    main()

# a = [1, 2, [3, 4, 5, 6, [7, [8, [9]]]]]
#
#
# def recursive(node, lst=[]):
#     for i in node:
#         if isinstance(i, list):
#             node = recursive(node=i, lst=lst)
#         else:
#             lst.append(i)
#     return lst
#
#
# print(recursive(node=a))
